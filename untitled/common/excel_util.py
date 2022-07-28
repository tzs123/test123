# coding:utf-8

"""
1、打开表单

2、读取标题 头部

3、读取所有的数据 (类型是？？)

4、指定单元格写入数据（使用静态方法，不要使用实例方法）
"""
import json

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHandler():
    """操作 Excel """

    def __init__(self, file):
        """初始化函数"""
        self.file = file
        self.wb = load_workbook(file)

    def open_sheet(self, name) -> Worksheet:
        """打开表单.
        在函数或者方法的后面 加 -> 类型：表示此函数返回值是一个 这样的类型
        函数注解。
        """
        wb = load_workbook(self.file)
        sheet = wb[name]
        wb.close()
        return sheet

    def header(self, sheet_name):
        """获取表单的表头"""
        sheet = self.open_sheet(sheet_name)
        headers = []
        for i in sheet[1]:
            headers.append(i.value)
        return headers

    def read(self, sheet_name):
        """读取所有的数据"""
        sheet = self.open_sheet(sheet_name)
        rows = list(sheet.rows)

        # 获取标题
        data = []
        for row in rows[1:]:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
                # 列表转成字典：要和 header 去 zip
                data_dict = dict(zip(self.header(sheet_name), row_data))
            data.append(data_dict)
        return data
    def ui_read(self,sheet_name):
        """读取所有的数据"""
        sheet = self.open_sheet(sheet_name)

        # 获取标题
        data = []
        for row in range(1,sheet.max_row+1):
            row_data = []
            for column in range(1,sheet.max_column+1):
                row_data.append(sheet.cell(row,column).value)
            data.append(row_data)
        return data
    @staticmethod
    def write(file, sheet_name, row, column, data):
        """写入 Excel 数据"""
        wb = load_workbook(str(file))
        sheet = wb[sheet_name]
        # 修改单元格
        sheet.cell(row, column).value = data
        # 保存
        wb.save(str(file))
        # 关闭
        wb.close()

excel = ExcelHandler(r'C:\Users\Administrator\.jenkins\workspace\psimsendemail\untitled\data\login_data.xlsx').read('app')
excel1 = ExcelHandler(r'C:\Users\Administrator\.jenkins\workspace\psimsendemail\untitled\data\register.xlsx').read('register')

print(excel[1]['是否执行'])


