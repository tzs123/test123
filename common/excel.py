# _*_ coding:utf-8 _*_

from openpyxl import Workbook,load_workbook


class ExcelUtil(object):

    def __init__(self, excel_path):
        self.wb = load_workbook(excel_path)
        self.template = """{"id":0,"url":"","case_name":"","header":"","method":"","body":"",
        "expect":"","actual":"","valiadate":""},"""  # 这个是写入用例的模板


    def read_excel(self):
        """读取excel，处理数据，并返回一个格式处理后的字典"""
        value = []
        for sheetname in self.wb.sheetnames:
            ws = self.wb[sheetname]
            cases_num = len(list(ws.values)) - 1  # 一个sheet中用例的数量
            case_list = list(ws.values)
            case_list.pop(0)  # 去掉表头
            cases_template = self.template * cases_num
            cases_template_list = eval("[" + cases_template[:-1] + "]")   # 与用例相同长度的模板

            for i in range(len(case_list)):  # i：第i个用例
                # 每个用例中字段是9个，因此这样写
                cases_template_list[i]['id'] = case_list[i][0]
                cases_template_list[i]['url'] = case_list[i][1]
                cases_template_list[i]['case_name'] = case_list[i][2]
                cases_template_list[i]['header'] = case_list[i][3]
                cases_template_list[i]['method'] = case_list[i][4]
                cases_template_list[i]['body'] = case_list[i][5]
                cases_template_list[i]['expect'] = case_list[i][6]
                cases_template_list[i]['actual'] = case_list[i][7]
                cases_template_list[i]['valiadate'] = case_list[i][8]

            value.append({"cases": cases_template_list})

        return value

    @staticmethod
    def write(file, sheet_name, row, column, data):
        """写入 Excel 数据"""
        wb = load_workbook(str(file))
        sheet = wb[sheet_name]
        # 修改单元格
        sheet.cell(row, column).value = data
        # 保存
        wb.save(str(file))
        # 关闭
        wb.close()